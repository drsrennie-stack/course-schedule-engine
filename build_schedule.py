#!/usr/bin/env python3
"""
Course Schedule Engine, build script.

One source of truth, many outputs. Edit schedule-data.xlsx (or schedule-data.json),
then run this script. It rebuilds:

  course-schedule.html        the interactive page, with fresh data baked in
  schedule-data.json          the canonical data file
  course-schedule.pdf         a print-ready schedule
  exports/*.ics               calendar file, due dates
  exports/*-spaced-repetition.csv   starter Anki deck, one card per topic
  exports/*-study-plan.txt    nightly tasks plus scheduled spaced reviews
  exports/*-planner-checklist.md    Markdown checklist for any task planner
  exports/*-slo-grade-map.csv every graded item mapped to an SLO

Usage:
  python build_schedule.py                 read the xlsx if present, else the json
  python build_schedule.py --source json   force the json as the source
  python build_schedule.py --source xlsx   force the xlsx as the source
  python build_schedule.py --no-pdf        skip the PDF (if reportlab is missing)

No network access, no student data, no external services.
Prepared for Dr. Sharilyn Rennie.
"""
import os
import re
import sys
import csv
import json
import io
import datetime as dt

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(BASE, "schedule-data.xlsx")
JSON = os.path.join(BASE, "schedule-data.json")
HTML = os.path.join(BASE, "course-schedule.html")
PDF = os.path.join(BASE, "course-schedule.pdf")
EXPORT_DIR = os.path.join(BASE, "exports")

ITEM_TYPES = {
    "rat": "Readiness Assurance",
    "application": "Application Activity",
    "lab-exam": "Lab Practical Exam",
    "midterm": "Midterm Exam",
    "final": "Final Exam",
}
EXAM_TYPES = {"lab-exam", "midterm", "final"}
SR_NOTE = "Review offsets, in days, after a topic is first taught."
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTHS_LONG = ["January", "February", "March", "April", "May", "June", "July",
               "August", "September", "October", "November", "December"]
WEEKDAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


# ----------------------------------------------------------------------
# Date helpers
# ----------------------------------------------------------------------
def parse_iso(s):
    return dt.date(*[int(x) for x in s.split("-")])


def norm_date(v):
    """Accept a string YYYY-MM-DD or a datetime and return YYYY-MM-DD."""
    if v is None:
        return ""
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return ""
    # tolerate slashes
    s = s.replace("/", "-")
    return s


def fmt_short(iso):
    d = parse_iso(iso)
    return WEEKDAYS[d.weekday()][:3] + ", " + MONTHS[d.month - 1] + " " + str(d.day)


def fmt_long(iso):
    d = parse_iso(iso)
    return MONTHS_LONG[d.month - 1] + " " + str(d.day) + ", " + str(d.year)


def fmt_range(a, b):
    da, db = parse_iso(a), parse_iso(b)
    if da.month == db.month:
        return "%s %d to %d" % (MONTHS[da.month - 1], da.day, db.day)
    return "%s %d to %s %d" % (MONTHS[da.month - 1], da.day, MONTHS[db.month - 1], db.day)


def add_days(iso, n):
    return (parse_iso(iso) + dt.timedelta(days=n)).strftime("%Y-%m-%d")


def weekday_name(iso):
    return WEEKDAYS[parse_iso(iso).weekday()]


# ----------------------------------------------------------------------
# Source readers
# ----------------------------------------------------------------------
def read_json(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def read_xlsx(path):
    """Reconstruct the schedule data dict from the spreadsheet template."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)

    # ---- Course sheet, field/value pairs in columns A and B ----
    cs = wb["Course"]
    fields = {}
    for row in cs.iter_rows(min_row=3, max_col=2, values_only=True):
        key, val = row[0], row[1]
        if key:
            fields[str(key).strip()] = val
    sample = str(fields.get("isSampleData", "FALSE")).strip().upper() in ("TRUE", "1", "YES")
    intervals = []
    for part in str(fields.get("srIntervals", "1, 3, 8, 21")).split(","):
        part = part.strip()
        if part:
            intervals.append(int(float(part)))

    course = {
        "code": str(fields.get("code", "")).strip(),
        "title": str(fields.get("title", "")).strip(),
        "institution": str(fields.get("institution", "")).strip(),
        "term": str(fields.get("term", "")).strip(),
        "instructor": str(fields.get("instructor", "")).strip(),
        "meetingPattern": str(fields.get("meetingPattern", "")).strip(),
        "format": str(fields.get("format", "")).strip(),
        "timezone": str(fields.get("timezone", "America/Los_Angeles")).strip(),
        "isSampleData": sample,
        "slos": [],
        "itemTypes": dict(ITEM_TYPES),
    }

    # ---- SLOs sheet ----
    for row in wb["SLOs"].iter_rows(min_row=3, max_col=2, values_only=True):
        code, desc = row[0], row[1]
        if code:
            course["slos"].append({"code": str(code).strip(),
                                   "description": str(desc or "").strip()})

    # ---- Schedule sheet, one row per class day ----
    days_by_date = {}
    week_meta = {}
    for row in wb["Schedule"].iter_rows(min_row=3, values_only=True):
        if row[0] in (None, ""):
            continue
        week = int(row[0])
        theme = str(row[1] or "").strip()
        wstart = norm_date(row[2])
        wend = norm_date(row[3])
        date = norm_date(row[4])
        if not date:
            continue
        weekday = str(row[5] or "").strip() or weekday_name(date)
        topic = str(row[6] or "").strip()
        focus = str(row[7] or "").strip()
        in_class = [str(c).strip() for c in row[8:12] if c not in (None, "")]
        nightly = [str(c).strip() for c in row[12:15] if c not in (None, "")]
        week_meta.setdefault(week, {"theme": theme, "start": wstart, "end": wend})
        days_by_date[date] = {
            "week": week, "date": date, "weekday": weekday, "topic": topic,
            "focus": focus, "inClass": in_class, "items": [], "nightly": nightly,
        }

    # ---- Graded Items sheet ----
    orphans = []
    for row in wb["Graded Items"].iter_rows(min_row=3, values_only=True):
        if row[0] in (None, ""):
            continue
        date = norm_date(row[1])
        item = {
            "id": str(row[0]).strip(),
            "title": str(row[2] or "").strip(),
            "type": str(row[3] or "").strip(),
            "points": int(float(row[4] or 0)),
            "due": norm_date(row[5]) or date,
            "slo": [s.strip() for s in re.split(r"[;,]", str(row[6] or "")) if s.strip()],
            "note": str(row[7] or "").strip(),
        }
        if date in days_by_date:
            days_by_date[date]["items"].append(item)
        else:
            orphans.append((item["id"], date))

    if orphans:
        for iid, date in orphans:
            warn("Graded item '%s' has Date '%s' which matches no row on the Schedule tab." % (iid, date))

    # ---- Assemble weeks ----
    by_week = {}
    for d in days_by_date.values():
        by_week.setdefault(d["week"], []).append(d)

    weeks = []
    for week in sorted(week_meta):
        meta = week_meta[week]
        wdays = sorted(by_week.get(week, []), key=lambda d: d["date"])
        for d in wdays:
            d.pop("week", None)
        if not meta["start"] and wdays:
            meta["start"] = wdays[0]["date"]
        if not meta["end"] and wdays:
            meta["end"] = wdays[-1]["date"]
        weeks.append({"week": week, "theme": meta["theme"],
                      "start": meta["start"], "end": meta["end"], "days": wdays})

    return {
        "schemaVersion": "1.0",
        "course": course,
        "weeks": weeks,
        "spacedRepetition": {"intervalsDays": intervals or [1, 3, 8, 21], "note": SR_NOTE},
    }


# ----------------------------------------------------------------------
# Validation and normalisation
# ----------------------------------------------------------------------
WARNINGS = []


def warn(msg):
    WARNINGS.append(msg)


def normalise(data):
    """Fill computed fields and run sanity checks."""
    course = data["course"]
    course.setdefault("itemTypes", dict(ITEM_TYPES))
    data.setdefault("spacedRepetition", {"intervalsDays": [1, 3, 8, 21], "note": SR_NOTE})

    slo_codes = {s["code"] for s in course.get("slos", [])}
    all_items = []
    for w in data["weeks"]:
        for day in w["days"]:
            day.setdefault("inClass", [])
            day.setdefault("nightly", [])
            day.setdefault("items", [])
            if not day.get("weekday") and day.get("date"):
                day["weekday"] = weekday_name(day["date"])
            for it in day["items"]:
                all_items.append(it)
                if it["type"] not in ITEM_TYPES:
                    warn("Item '%s' has type '%s'. Expected one of: %s."
                         % (it["id"], it["type"], ", ".join(ITEM_TYPES)))
                for code in it.get("slo", []):
                    if slo_codes and code not in slo_codes:
                        warn("Item '%s' references SLO '%s' which is not defined on the SLOs tab."
                             % (it["id"], code))
                try:
                    parse_iso(it["due"])
                except Exception:
                    warn("Item '%s' has an unreadable due date: '%s'." % (it["id"], it.get("due")))

    course["totalPoints"] = sum(int(i.get("points", 0)) for i in all_items)
    return data


# ----------------------------------------------------------------------
# Writers, JSON and HTML
# ----------------------------------------------------------------------
def write_json(data, path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
        f.write("\n")


def update_html(data, path):
    """Replace the embedded schedule-data block inside the HTML page."""
    if not os.path.exists(path):
        warn("course-schedule.html not found, skipped the HTML refresh.")
        return False
    with open(path, encoding="utf-8") as f:
        html = f.read()
    payload = json.dumps(data, indent=2, ensure_ascii=False)
    pattern = re.compile(
        r'(<script id="schedule-data" type="application/json">)(.*?)(</script>)',
        re.DOTALL)
    if not pattern.search(html):
        warn("Could not find the schedule-data block in the HTML, skipped.")
        return False
    new_html = pattern.sub(lambda m: m.group(1) + "\n" + payload + "\n" + m.group(3), html)
    with open(path, "w", encoding="utf-8") as f:
        f.write(new_html)
    return True


# ----------------------------------------------------------------------
# Export builders, these mirror the in-browser export buttons
# ----------------------------------------------------------------------
def iter_items(data):
    """Yield graded items in due-date order, each with its day context."""
    rows = []
    for w in data["weeks"]:
        for day in w["days"]:
            for it in day["items"]:
                rows.append({"item": it, "date": day["date"], "week": w["week"],
                             "weekTheme": w["theme"], "topic": day["topic"]})
    rows.sort(key=lambda r: r["date"])
    return rows


def csv_string(rows):
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\r\n")
    for r in rows:
        w.writerow(r)
    return buf.getvalue()


def ics_escape(s):
    return (str(s).replace("\\", "\\\\").replace(";", "\\;")
            .replace(",", "\\,").replace("\n", "\\n"))


def ics_fold(line):
    if len(line) <= 73:
        return line
    out, i = "", 0
    while i < len(line):
        size = 73 if i == 0 else 72
        chunk = line[i:i + size]
        out += ("" if i == 0 else "\r\n ") + chunk
        i += size
    return out


def build_ics(data):
    course = data["course"]
    stamp = dt.datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    L = ["BEGIN:VCALENDAR", "VERSION:2.0",
         "PRODID:-//MedMasters Collaborative//Course Schedule Engine//EN",
         "CALSCALE:GREGORIAN", "METHOD:PUBLISH",
         "X-WR-CALNAME:" + ics_escape("%s %s due dates" % (course["code"], course["title"]))]
    for r in iter_items(data):
        it = r["item"]
        d = it["due"].replace("-", "")
        d_end = add_days(it["due"], 1).replace("-", "")
        label = ITEM_TYPES.get(it["type"], it["type"])
        desc = "%s. Worth %d points. %sSLOs: %s." % (
            label, it["points"], (it["note"] + " ") if it.get("note") else "",
            ", ".join(it.get("slo", [])))
        L += ["BEGIN:VEVENT",
              "UID:%s-%s@course-schedule-engine" % (it["id"], d),
              "DTSTAMP:" + stamp,
              "DTSTART;VALUE=DATE:" + d,
              "DTEND;VALUE=DATE:" + d_end,
              ics_fold("SUMMARY:" + ics_escape("%s: %s (%d pts)" % (course["code"], it["title"], it["points"]))),
              ics_fold("DESCRIPTION:" + ics_escape(desc)),
              "CATEGORIES:" + ics_escape(label),
              "TRANSP:TRANSPARENT"]
        if it["type"] in EXAM_TYPES:
            L += ["BEGIN:VALARM", "ACTION:DISPLAY", "TRIGGER:-P1D",
                  ics_fold("DESCRIPTION:" + ics_escape("Tomorrow: " + it["title"])), "END:VALARM"]
        L.append("END:VEVENT")
    L.append("END:VCALENDAR")
    return "\r\n".join(L) + "\r\n"


def slug(s):
    return re.sub(r"^-|-$", "", re.sub(r"[^a-z0-9]+", "-", s.lower()))


def build_anki(data):
    course = data["course"]
    lines = ["#separator:Comma", "#html:false", "#columns:Front Back Deck Tags"]
    rows = []
    code = course["code"].replace(" ", "")
    for w in data["weeks"]:
        deck = "%s::Week %d %s" % (code, w["week"], w["theme"].replace(":", " "))
        for day in w["days"]:
            if any(i["type"] in EXAM_TYPES for i in day["items"]):
                continue
            rows.append(["Describe the anatomy of: " + day["topic"],
                         day.get("focus", ""),
                         deck,
                         "BIO004 week%d %s" % (w["week"], slug(day["topic"]))])
    return "\n".join(lines) + "\n" + csv_string(rows).replace("\r\n", "\n")


def build_study_plan(data):
    course = data["course"]
    intervals = data["spacedRepetition"]["intervalsDays"]
    course_end = data["weeks"][-1]["end"]
    review_by_date = {}
    for w in data["weeks"]:
        for day in w["days"]:
            if any(i["type"] in EXAM_TYPES for i in day["items"]):
                continue
            for off in intervals:
                rd = add_days(day["date"], off)
                if rd > course_end:
                    continue
                review_by_date.setdefault(rd, []).append(day["topic"])

    L = ["%s %s, Nightly Study Plan" % (course["code"], course["title"]),
         "%s. Prepared by Dr. Sharilyn Rennie." % course["term"],
         "Recommended tasks are not graded. Spaced reviews use offsets of %s days."
         % ", ".join(str(x) for x in intervals),
         "Paste this into any planner or spaced-repetition app.", ""]
    for w in data["weeks"]:
        L += ["==============================",
              "WEEK %d: %s  (%s)" % (w["week"], w["theme"], fmt_range(w["start"], w["end"])),
              "=============================="]
        for day in w["days"]:
            L.append("")
            L.append("%s, %s  ::  %s" % (day["weekday"], fmt_long(day["date"]), day["topic"]))
            for it in day["items"]:
                L.append("  GRADED, due today: %s (%d pts)" % (it["title"], it["points"]))
            L.append("  Recommended tonight:")
            for n in day["nightly"]:
                L.append("    [ ] " + n)
            rev = review_by_date.get(day["date"])
            if rev:
                seen = []
                for t in rev:
                    if t not in seen:
                        seen.append(t)
                L.append("  Spaced review tonight:")
                for t in seen:
                    L.append("    [ ] Re-test yourself on: " + t)
        L.append("")
    return "\n".join(L) + "\n"


def build_checklist(data):
    course = data["course"]
    total = course["totalPoints"]
    n_items = sum(len(d["items"]) for w in data["weeks"] for d in w["days"])
    L = ["# %s %s, Course Checklist" % (course["code"], course["title"]), "",
         "_%s. %d graded items, %d points. Prepared by Dr. Sharilyn Rennie._"
         % (course["term"], n_items, total), ""]
    for w in data["weeks"]:
        L.append("## Week %d: %s  (%s)" % (w["week"], w["theme"], fmt_range(w["start"], w["end"])))
        L.append("")
        for day in w["days"]:
            L.append("### %s, %s" % (fmt_short(day["date"]), day["topic"]))
            L.append("")
            for it in day["items"]:
                L.append("- [ ] **DUE %s, %s:** %s (%d pts)"
                         % (fmt_short(it["due"]), ITEM_TYPES.get(it["type"], it["type"]),
                            it["title"], it["points"]))
            if day["nightly"]:
                L.append("- Recommended tonight, not graded:")
                for n in day["nightly"]:
                    L.append("  - [ ] " + n)
            L.append("")
    return "\n".join(L) + "\n"


def build_slo_csv(data):
    course = data["course"]
    total = course["totalPoints"] or 1
    slo_map = {s["code"]: s["description"] for s in course.get("slos", [])}
    rows = [["Item ID", "Item", "Type", "Week", "Due Date", "Points",
             "Percent of Course", "SLO Codes", "SLO Descriptions"]]
    for r in iter_items(data):
        it = r["item"]
        codes = it.get("slo", [])
        rows.append([it["id"], it["title"], ITEM_TYPES.get(it["type"], it["type"]),
                     "Week %d" % r["week"], it["due"], it["points"],
                     "%.1f%%" % (it["points"] / total * 100),
                     "; ".join(codes),
                     "; ".join(slo_map.get(c, c) for c in codes)])
    rows.append([])
    rows.append(["SLO Summary"])
    rows.append(["SLO Code", "SLO Description", "Items Touching SLO", "Points Touching SLO"])
    for s in course.get("slos", []):
        hits = [r for r in iter_items(data) if s["code"] in r["item"].get("slo", [])]
        rows.append([s["code"], s["description"], len(hits),
                     sum(r["item"]["points"] for r in hits)])
    return csv_string(rows)


# ----------------------------------------------------------------------
# PDF builder
# ----------------------------------------------------------------------
def build_pdf(data, path):
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from reportlab.lib.colors import HexColor, white
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_LEFT
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                        TableStyle, KeepTogether, HRFlowable)
    except ImportError:
        warn("reportlab is not installed, skipped the PDF. Install it with: pip install reportlab")
        return False

    NAVY = HexColor("#1E3D4C")
    GOLD = HexColor("#B8924A")
    TERRA = HexColor("#A0522D")
    GRAY = HexColor("#6B7378")
    LINE = HexColor("#C7CDD0")
    TINT = HexColor("#EDF1F3")

    course = data["course"]

    def st(name, **kw):
        base = dict(fontName="Helvetica", fontSize=9.5, leading=13, textColor=NAVY,
                    alignment=TA_LEFT, spaceBefore=0, spaceAfter=0)
        base.update(kw)
        return ParagraphStyle(name, **base)

    s_eyebrow = st("eyebrow", fontName="Helvetica-Bold", fontSize=8, textColor=TERRA, leading=11)
    s_h1 = st("h1", fontName="Helvetica-Bold", fontSize=22, leading=26)
    s_sub = st("sub", fontName="Helvetica-Bold", fontSize=12, textColor=TERRA, leading=15)
    s_meta = st("meta", fontName="Helvetica-Oblique", fontSize=9, textColor=GRAY, leading=12)
    s_week = st("week", fontName="Helvetica-Bold", fontSize=12, textColor=white, leading=15)
    s_topic = st("topic", fontName="Helvetica-Bold", fontSize=10.5, leading=13)
    s_date = st("date", fontName="Helvetica-Bold", fontSize=8, textColor=TERRA, leading=11)
    s_focus = st("focus", fontName="Helvetica-Oblique", fontSize=9, textColor=GRAY, leading=12)
    s_body = st("body", fontSize=9, leading=12)
    s_grade = st("grade", fontName="Helvetica-Bold", fontSize=9, leading=12)
    s_exam = st("exam", fontName="Helvetica-Bold", fontSize=9, textColor=TERRA, leading=12)
    s_night = st("night", fontSize=8.5, textColor=NAVY, leading=11.5)
    s_nlabel = st("nlabel", fontName="Helvetica-Oblique", fontSize=8, textColor=GRAY, leading=11)

    story = []

    # ---- Header ----
    story.append(Paragraph("%s &nbsp;&middot;&nbsp; %s &nbsp;&middot;&nbsp; %s"
                           % (course["code"], course["title"], course["institution"]), s_eyebrow))
    story.append(Spacer(1, 4))
    story.append(Paragraph("Course Schedule", s_h1))
    story.append(Spacer(1, 2))
    story.append(Paragraph("%s &nbsp;&middot;&nbsp; %s" % (course["title"], course["term"]), s_sub))
    story.append(Spacer(1, 4))
    story.append(Paragraph("%s. Meets %s." % (course["format"], course["meetingPattern"]), s_meta))
    if course.get("isSampleData"):
        story.append(Spacer(1, 6))
        flag = Table([[Paragraph("SAMPLE DATA, replace dates and topics with your real course",
                                 st("flag", fontName="Helvetica-Bold", fontSize=8, textColor=NAVY))]],
                     colWidths=[3.5 * inch])
        flag.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), TINT),
                                  ("BOX", (0, 0), (-1, -1), 0.75, NAVY),
                                  ("TOPPADDING", (0, 0), (-1, -1), 4),
                                  ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                                  ("LEFTPADDING", (0, 0), (-1, -1), 8)]))
        story.append(flag)
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=1, color=LINE))
    story.append(Spacer(1, 12))

    # ---- Graded items summary ----
    story.append(Paragraph("Graded Items and Due Dates", s_sub))
    story.append(Spacer(1, 6))
    head = ["Due Date", "Item", "Type", "Points"]
    body = [head]
    rows = iter_items(data)
    for r in rows:
        it = r["item"]
        body.append([fmt_short(it["due"]), it["title"],
                     ITEM_TYPES.get(it["type"], it["type"]), str(it["points"])])
    body.append(["", "Course total", "", str(course["totalPoints"])])
    tbl = Table(body, colWidths=[1.05 * inch, 3.15 * inch, 1.55 * inch, 0.75 * inch], repeatRows=1)
    ts = [("BACKGROUND", (0, 0), (-1, 0), NAVY),
          ("TEXTCOLOR", (0, 0), (-1, 0), white),
          ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
          ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
          ("FONTSIZE", (0, 0), (-1, -1), 8.5),
          ("TEXTCOLOR", (0, 1), (-1, -1), NAVY),
          ("ALIGN", (3, 0), (3, -1), "CENTER"),
          ("ROWBACKGROUNDS", (0, 1), (-1, -2), [white, HexColor("#FAFAF9")]),
          ("BACKGROUND", (0, -1), (-1, -1), TINT),
          ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
          ("LINEBELOW", (0, 0), (-1, -1), 0.5, LINE),
          ("BOX", (0, 0), (-1, -1), 0.75, NAVY),
          ("TOPPADDING", (0, 0), (-1, -1), 5),
          ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
          ("LEFTPADDING", (0, 0), (-1, -1), 7)]
    for i, r in enumerate(rows, start=1):
        if r["item"]["type"] in EXAM_TYPES:
            ts.append(("LINEBEFORE", (0, i), (0, i), 3, GOLD))
            ts.append(("TEXTCOLOR", (0, i), (-1, i), TERRA))
    tbl.setStyle(TableStyle(ts))
    story.append(tbl)
    story.append(Spacer(1, 16))

    # ---- Per-week schedule ----
    for w in data["weeks"]:
        block = []
        wk = Table([[Paragraph("Week %d: %s" % (w["week"], w["theme"]), s_week),
                     Paragraph(fmt_range(w["start"], w["end"]),
                               st("wkdate", fontName="Helvetica-Bold", fontSize=8.5,
                                  textColor=white, leading=15, alignment=2))]],
                   colWidths=[5.0 * inch, 1.5 * inch])
        wk.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), NAVY),
                                ("TOPPADDING", (0, 0), (-1, -1), 6),
                                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                                ("LEFTPADDING", (0, 0), (-1, -1), 9),
                                ("RIGHTPADDING", (0, 0), (-1, -1), 9),
                                ("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        block.append(wk)
        block.append(Spacer(1, 6))

        for di, day in enumerate(w["days"]):
            day_flow = []
            is_exam = any(i["type"] in EXAM_TYPES for i in day["items"])
            day_flow.append(Paragraph("%s &nbsp;&middot;&nbsp; %s"
                                      % (day["weekday"], fmt_short(day["date"])), s_date))
            day_flow.append(Paragraph(day["topic"], s_topic))
            if day.get("focus"):
                day_flow.append(Paragraph(day["focus"], s_focus))
            if day.get("inClass"):
                day_flow.append(Spacer(1, 2))
                day_flow.append(Paragraph("In class: " + "; ".join(day["inClass"]), s_body))
            for it in day["items"]:
                style = s_exam if it["type"] in EXAM_TYPES else s_grade
                label = ITEM_TYPES.get(it["type"], it["type"])
                day_flow.append(Spacer(1, 2))
                day_flow.append(Paragraph(
                    "DUE %s &nbsp; %s: %s &nbsp; (%d pts)"
                    % (fmt_short(it["due"]), label, it["title"], it["points"]), style))
            if day["nightly"]:
                day_flow.append(Spacer(1, 3))
                day_flow.append(Paragraph("Recommended tonight, not graded:", s_nlabel))
                for n in day["nightly"]:
                    day_flow.append(Paragraph("&bull;&nbsp; " + n, s_night))

            cell_bg = TINT if is_exam else white
            card = Table([[day_flow]], colWidths=[6.5 * inch])
            card.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), white),
                ("BOX", (0, 0), (-1, -1), 0.75, LINE),
                ("LINEBEFORE", (0, 0), (0, 0), 4, GOLD if is_exam else NAVY),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ("LEFTPADDING", (0, 0), (-1, -1), 9),
                ("RIGHTPADDING", (0, 0), (-1, -1), 9)]))
            block.append(KeepTogether([card, Spacer(1, 6)]))

        story.append(KeepTogether(block[:2]))
        for fl in block[2:]:
            story.append(fl)
        story.append(Spacer(1, 10))

    # ---- Footer on every page ----
    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(GRAY)
        canvas.drawString(0.75 * inch, 0.5 * inch,
                          "%s %s  |  %s  |  Prepared by Dr. Sharilyn Rennie"
                          % (course["code"], course["title"], course["term"]))
        canvas.drawRightString(7.75 * inch, 0.5 * inch, "Page %d" % doc.page)
        canvas.setStrokeColor(LINE)
        canvas.setLineWidth(0.5)
        canvas.line(0.75 * inch, 0.62 * inch, 7.75 * inch, 0.62 * inch)
        canvas.restoreState()

    doc = SimpleDocTemplate(path, pagesize=letter,
                            leftMargin=0.75 * inch, rightMargin=0.75 * inch,
                            topMargin=0.7 * inch, bottomMargin=0.8 * inch,
                            title="Course Schedule, %s %s" % (course["code"], course["title"]),
                            author="Dr. Sharilyn Rennie")
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    return True


# ----------------------------------------------------------------------
# Orchestration
# ----------------------------------------------------------------------
def main():
    args = sys.argv[1:]
    source = "auto"
    do_pdf = True
    if "--source" in args:
        source = args[args.index("--source") + 1]
    if "--no-pdf" in args:
        do_pdf = False
    if "--help" in args or "-h" in args:
        print(__doc__)
        return

    # ---- Choose the source ----
    if source == "json":
        path = JSON
    elif source == "xlsx":
        path = XLSX
    else:
        path = XLSX if os.path.exists(XLSX) else JSON

    if not os.path.exists(path):
        print("ERROR: no source file found. Expected schedule-data.xlsx or schedule-data.json.")
        sys.exit(1)

    print("Course Schedule Engine")
    print("Source: %s" % os.path.basename(path))
    data = read_xlsx(path) if path.endswith(".xlsx") else read_json(path)
    data = normalise(data)

    course = data["course"]
    n_items = sum(len(d["items"]) for w in data["weeks"] for d in w["days"])
    print("Loaded: %s %s, %d weeks, %d graded items, %d points."
          % (course["code"], course["title"], len(data["weeks"]), n_items, course["totalPoints"]))

    # ---- Canonical JSON ----
    write_json(data, JSON)
    print("  wrote schedule-data.json")

    # ---- HTML refresh ----
    if update_html(data, HTML):
        print("  refreshed course-schedule.html")

    # ---- Exports ----
    os.makedirs(EXPORT_DIR, exist_ok=True)
    base = course["code"].replace(" ", "-").lower()
    outputs = {
        base + "-schedule.ics": build_ics(data),
        base + "-spaced-repetition.csv": build_anki(data),
        base + "-study-plan.txt": build_study_plan(data),
        base + "-planner-checklist.md": build_checklist(data),
        base + "-slo-grade-map.csv": build_slo_csv(data),
    }
    for name, text in outputs.items():
        with open(os.path.join(EXPORT_DIR, name), "w", encoding="utf-8", newline="") as f:
            f.write(text)
        print("  wrote exports/%s" % name)

    # ---- PDF ----
    if do_pdf:
        if build_pdf(data, PDF):
            print("  wrote course-schedule.pdf")

    # ---- Warnings ----
    if WARNINGS:
        print("\nChecks flagged %d item(s) to review:" % len(WARNINGS))
        for w in WARNINGS:
            print("  - " + w)
    else:
        print("\nAll checks passed.")
    print("Done.")


if __name__ == "__main__":
    main()
